import csv
import json
from datetime import date, datetime, time
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import (
    Http404,
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.safestring import mark_safe
from django.views import View
from django.views.decorators.http import require_POST
from django.views.generic import DetailView, ListView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from acarreapp.tenancy import get_current_empresa
from empresa.models import Cliente, Vehiculo
from notificaciones.utils import send_webpush_to_empresa
from servicios.models import Servicio

from .forms import RutaForm
from .models import CierreRuta, MovimientoCaja, Ruta
from .services import calcular_cierre_ruta, cerrar_ruta


def is_gerente(user):
    role = getattr(getattr(user, "userprofile", None), "rol", "")
    return user.is_superuser or user.is_staff or role == "GERENTE"


def is_conductor(user):
    role = getattr(getattr(user, "userprofile", None), "rol", "")
    return role == "CONDUCTOR" and not (user.is_staff or user.is_superuser)


def _empresa_or_404():
    empresa = get_current_empresa()
    if empresa is None:
        raise Http404("No hay empresa activa.")
    return empresa


def _ruta_qs():
    empresa = _empresa_or_404()
    return Ruta.objects.filter(empresa=empresa).select_related("vehiculo", "conductor", "empresa")


def _ruta_autorizada(pk, user, *, gerente_only=False):
    ruta = get_object_or_404(_ruta_qs(), pk=pk)
    if gerente_only and not is_gerente(user):
        return None
    if is_conductor(user) and ruta.conductor_id != user.id:
        return None
    return ruta


def _parse_positive_int(value):
    try:
        parsed = int(value or "0")
    except (TypeError, ValueError):
        return 0
    return parsed if parsed > 0 else 0


def _safe_export(value):
    if value is None:
        return ""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _xls(value):
    value = _safe_export(value)
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    return str(value)


def _money_fmt(cell):
    cell.number_format = u'[$$-409] #,##0'
    return cell


def _auto_fit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


def _cierre_para_ruta(ruta, user=None):
    return CierreRuta.objects.filter(ruta=ruta, empresa=ruta.empresa).first() or calcular_cierre_ruta(ruta, user)


def _cierre_context(ruta, cierre):
    servicios = ruta.servicios.select_related("cliente").prefetch_related("comentarios").order_by("orden", "id")
    servicios_list = list(servicios)
    total_venta = sum((s.valor or 0) for s in servicios_list)
    total_cobrado = int(getattr(cierre, "total_cobrado", 0) or 0)
    pendiente_cobro = int(getattr(cierre, "total_pendiente", 0) or 0)
    base_efectivo = int(ruta.base_efectivo or 0)
    total_ingresos = int(getattr(cierre, "total_ingresos", 0) or 0)
    total_gastos = int(getattr(cierre, "total_gastos", 0) or 0)
    ingresos_en_ruta = max(total_ingresos - base_efectivo, 0)
    efectivo_entregar = base_efectivo + ingresos_en_ruta - total_gastos
    utilidad_operativa = total_venta - total_gastos
    return {
        "ruta": ruta,
        "cierre": cierre,
        "servicios": servicios_list,
        "total_venta": total_venta,
        "total_cobrado": total_cobrado,
        "pendiente_cobro": pendiente_cobro,
        "base_efectivo": base_efectivo,
        "ingresos_en_ruta": ingresos_en_ruta,
        "total_gastos": total_gastos,
        "efectivo_entregar": efectivo_entregar,
        "utilidad_operativa": utilidad_operativa,
    }


@method_decorator(login_required, name="dispatch")
class RutasListView(ListView):
    model = Ruta
    template_name = "rutas/list.html"
    context_object_name = "rutas"
    paginate_by = 25

    def get_queryset(self):
        qs = _ruta_qs()
        if is_conductor(self.request.user):
            qs = qs.filter(conductor=self.request.user, estado="ACTIVA")

        params = self.request.GET
        if not is_conductor(self.request.user):
            if params.get("activas") == "1":
                qs = qs.filter(estado="ACTIVA")
            elif params.get("cerradas") == "1":
                qs = qs.filter(estado="CERRADA")

        desde = (params.get("desde") or "").strip()
        hasta = (params.get("hasta") or "").strip()
        if desde:
            qs = qs.filter(fecha_salida__gte=desde)
        if hasta:
            qs = qs.filter(fecha_salida__lte=hasta)

        vehiculos_ids = [v for v in params.getlist("vehiculos") if v]
        if vehiculos_ids:
            qs = qs.filter(vehiculo_id__in=vehiculos_ids)

        clientes_ids = [c for c in params.getlist("clientes") if c]
        if clientes_ids:
            qs = qs.filter(servicios__cliente_id__in=clientes_ids)

        q = (params.get("q") or "").strip()
        if q:
            base = (
                Q(nombre__icontains=q)
                | Q(vehiculo__placa__icontains=q)
                | Q(conductor__username__icontains=q)
                | Q(conductor__first_name__icontains=q)
                | Q(conductor__last_name__icontains=q)
                | Q(estado__icontains=q)
            )
            if q.isdigit():
                base |= Q(id=int(q))
            qs = qs.filter(
                base
                | Q(servicios__cliente__nombre__icontains=q)
                | Q(servicios__origen__icontains=q)
                | Q(servicios__destino__icontains=q)
            )

        return qs.distinct().order_by("-created_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["es_gerente"] = is_gerente(self.request.user)
        ctx["es_conductor"] = is_conductor(self.request.user)
        return ctx


@method_decorator(login_required, name="dispatch")
class RutaDetailView(DetailView):
    model = Ruta
    template_name = "rutas/detail.html"
    context_object_name = "ruta"

    def get_queryset(self):
        qs = _ruta_qs()
        if is_conductor(self.request.user):
            qs = qs.filter(conductor=self.request.user)
        return qs

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if is_conductor(request.user) and self.object.estado != "ACTIVA":
            messages.warning(request, "Esta ruta ya fue cerrada.")
            return redirect("rutas:list")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ruta = self.object
        servicios = ruta.servicios.select_related("cliente").order_by("orden", "id")
        servicios_list = list(servicios)
        movs = ruta.movimientos.all().order_by("-timestamp")
        movs_list = list(movs)

        valor_total = sum(s.valor for s in servicios_list)
        total_cobrado = sum(s.total_pagado for s in servicios_list)
        total_pendiente = sum(s.saldo_cartera for s in servicios_list)
        total_gastos = sum(m.valor for m in movs_list if m.tipo == "GASTO")
        total_ingresos = int(ruta.base_efectivo or 0) + sum(m.valor for m in movs_list if m.tipo == "INGRESO")
        disponible = total_ingresos - total_gastos

        ctx.update(
            {
                "servicios": servicios_list,
                "movimientos": movs_list,
                "tot_servicios": len(servicios_list),
                "valor_total": valor_total,
                "total_cobrado": total_cobrado,
                "total_pendiente": total_pendiente,
                "total_gastos": total_gastos,
                "total_ingresos": total_ingresos,
                "disponible": disponible,
                "utilidad_neta": total_cobrado - total_gastos,
                "ingresos_sin_base": max(total_ingresos - int(ruta.base_efectivo or 0), 0),
                "es_gerente": is_gerente(self.request.user),
                "es_conductor": is_conductor(self.request.user) and ruta.conductor_id == self.request.user.id,
                "add_servicio_url": reverse("servicios:crear") + f"?ruta={ruta.pk}",
            }
        )
        return ctx


@login_required
@user_passes_test(is_gerente)
def crear_ruta(request):
    if get_current_empresa() is None:
        return HttpResponseForbidden("No hay empresa activa.")
    if request.method == "POST":
        form = RutaForm(request.POST)
        if form.is_valid():
            ruta = form.save()
            messages.success(request, f"Ruta #{ruta.pk} creada.")
            return redirect("servicios:por_ruta", ruta_id=ruta.pk)
    else:
        form = RutaForm()
    return render(request, "rutas/crear_ruta.html", {"form": form})


@login_required
@user_passes_test(is_gerente)
@require_POST
def borrar_ruta(request, pk):
    ruta = _ruta_autorizada(pk, request.user, gerente_only=True)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    if ruta.estado != "ACTIVA":
        messages.error(request, "No puedes borrar una ruta cerrada.")
        return redirect("rutas:hoja", pk=ruta.pk)
    ruta.delete()
    messages.success(request, "Ruta eliminada.")
    return redirect("rutas:list")


@login_required
@user_passes_test(is_gerente)
@require_POST
def cerrar_ruta_view(request, pk):
    ruta = _ruta_autorizada(pk, request.user, gerente_only=True)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    try:
        cierre = cerrar_ruta(ruta, request.user)
    except ValueError as exc:
        messages.error(request, f"No se pudo cerrar la ruta: {exc}")
        return redirect("rutas:hoja", pk=ruta.pk)
    except IntegrityError:
        messages.error(request, "No se pudo cerrar la ruta por integridad de datos.")
        return redirect("rutas:hoja", pk=ruta.pk)
    messages.success(request, f"Ruta #{ruta.pk} cerrada.")
    return redirect("rutas:cierre_resumen", ruta_id=cierre.ruta_id)


@login_required
@require_POST
def agregar_gasto(request, pk):
    ruta = _ruta_autorizada(pk, request.user)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    if ruta.estado != "ACTIVA":
        messages.error(request, "La ruta esta cerrada.")
        return redirect("rutas:hoja", pk=pk)
    valor = _parse_positive_int(request.POST.get("valor"))
    concepto = (request.POST.get("concepto") or "").strip()
    if not valor:
        messages.error(request, "El valor debe ser positivo.")
    else:
        MovimientoCaja.objects.create(
            empresa=ruta.empresa,
            ruta=ruta,
            tipo="GASTO",
            concepto=concepto or "Gasto",
            valor=valor,
            usuario=request.user,
        )
        messages.success(request, "Gasto registrado.")
    return redirect("rutas:hoja", pk=ruta.id)


@login_required
@require_POST
def agregar_ingreso_extra(request, pk):
    ruta = _ruta_autorizada(pk, request.user)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    if ruta.estado != "ACTIVA":
        messages.error(request, "La ruta esta cerrada.")
        return redirect("rutas:hoja", pk=pk)
    valor = _parse_positive_int(request.POST.get("valor"))
    concepto = (request.POST.get("concepto") or "").strip()
    if not valor:
        messages.error(request, "El valor debe ser positivo.")
    else:
        MovimientoCaja.objects.create(
            empresa=ruta.empresa,
            ruta=ruta,
            tipo="INGRESO",
            concepto=concepto or "Ingreso extra",
            valor=valor,
            usuario=request.user,
        )
        messages.success(request, "Ingreso registrado.")
    return redirect("rutas:hoja", pk=ruta.id)


@login_required
@user_passes_test(is_gerente)
def cierre_resumen(request, ruta_id: int):
    ruta = _ruta_autorizada(ruta_id, request.user, gerente_only=True)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    cierre = _cierre_para_ruta(ruta, request.user)
    return render(request, "rutas/cierre_resumen.html", _cierre_context(ruta, cierre))


@login_required
def recorrido_ruta_view(request, ruta_id: int):
    ruta = _ruta_autorizada(ruta_id, request.user)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")

    puntos = []
    for servicio in ruta.servicios.all().order_by("orden", "id"):
        if servicio.recogido_en and servicio.lat_recogida is not None and servicio.lon_recogida is not None:
            puntos.append(
                {
                    "ts": servicio.recogido_en.isoformat(),
                    "lat": float(servicio.lat_recogida),
                    "lon": float(servicio.lon_recogida),
                    "tipo": "recogida",
                    "label": f"Recogida - Serv #{servicio.id}",
                }
            )
        if servicio.entregado_en and servicio.lat_entrega is not None and servicio.lon_entrega is not None:
            puntos.append(
                {
                    "ts": servicio.entregado_en.isoformat(),
                    "lat": float(servicio.lat_entrega),
                    "lon": float(servicio.lon_entrega),
                    "tipo": "entrega",
                    "label": f"Entrega - Serv #{servicio.id}",
                }
            )

    puntos.sort(key=lambda p: p["ts"])
    return render(request, "rutas/recorrido.html", {"ruta": ruta, "puntos_json": mark_safe(json.dumps(puntos))})


@login_required
@user_passes_test(is_gerente)
def exportar_cierre_csv(request, ruta_id: int):
    ruta = _ruta_autorizada(ruta_id, request.user, gerente_only=True)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    cierre = _cierre_para_ruta(ruta, request.user)
    servicios = ruta.servicios.select_related("cliente").order_by("orden", "id")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="cierre_ruta_{ruta.id}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Resumen de Cierre"])
    writer.writerow(["Ruta", _safe_export(str(ruta))])
    writer.writerow(["Total servicios", cierre.total_servicios])
    writer.writerow(["Cobrado", cierre.total_cobrado])
    writer.writerow(["Pendiente", cierre.total_pendiente])
    writer.writerow(["Ingresos", cierre.total_ingresos])
    writer.writerow(["Gastos", cierre.total_gastos])
    writer.writerow(["Utilidad neta", cierre.utilidad_neta])
    writer.writerow([])
    writer.writerow(["ID", "Cliente", "Origen", "Destino", "Valor", "Pagado", "Saldo", "Estado pago"])
    for servicio in servicios:
        writer.writerow(
            [
                servicio.id,
                _safe_export(getattr(servicio.cliente, "nombre", "")),
                _safe_export(servicio.origen),
                _safe_export(servicio.destino),
                servicio.valor,
                servicio.total_pagado,
                servicio.saldo_cartera,
                _safe_export(servicio.get_estado_pago_display()),
            ]
        )
    return response


@login_required
@user_passes_test(is_gerente)
def exportar_cierre_xlsx(request, ruta_id: int):
    ruta = _ruta_autorizada(ruta_id, request.user, gerente_only=True)
    if ruta is None:
        return HttpResponseForbidden("No autorizado")
    cierre = _cierre_para_ruta(ruta, request.user)
    ctx = _cierre_context(ruta, cierre)
    servicios = ctx["servicios"]

    header_fill = PatternFill("solid", fgColor="1F2A5A")
    header_font = Font(color="FFFFFF", bold=True)
    subhead_font = Font(color="1F2A5A", bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E6E8F0"),
        right=Side(style="thin", color="E6E8F0"),
        top=Side(style="thin", color="E6E8F0"),
        bottom=Side(style="thin", color="E6E8F0"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:F1")
    ws["A1"] = _xls(f"Cierre de ruta - {ruta.nombre or '(sin nombre)'} #{ruta.id}")
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    resumen = [
        ("Estado", (ruta.estado or "").title()),
        ("Vehiculo", getattr(ruta.vehiculo, "placa", ruta.vehiculo)),
        ("Conductor", getattr(ruta.conductor, "username", ruta.conductor)),
        ("Salida", ruta.fecha_salida),
    ]
    for row, (label, value) in enumerate(resumen, start=3):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = bold
        ws[f"B{row}"] = _xls(value)

    metrics = [
        ("Valor total de servicios", ctx["total_venta"]),
        ("Cobrado total", ctx["total_cobrado"]),
        ("Pendiente por cobrar", ctx["pendiente_cobro"]),
        ("Utilidad operativa", ctx["utilidad_operativa"]),
    ]
    for row, (label, value) in enumerate(metrics, start=3):
        ws[f"D{row}"] = label
        ws[f"D{row}"].font = subhead_font
        ws[f"E{row}"] = _xls(value)
        _money_fmt(ws[f"E{row}"])

    ws["A8"] = "Caja de ruta"
    ws["A8"].font = subhead_font
    for row, (label, value) in enumerate(
        [
            ("Base", ctx["base_efectivo"]),
            ("Ingresos en ruta", ctx["ingresos_en_ruta"]),
            ("Gastos", -abs(ctx["total_gastos"])),
            ("Efectivo a entregar", ctx["efectivo_entregar"]),
        ],
        start=9,
    ):
        ws.cell(row=row, column=2, value=label).font = bold
        ws.cell(row=row, column=3, value=_xls(value))
        _money_fmt(ws.cell(row=row, column=3)).alignment = right
        for col in range(2, 4):
            ws.cell(row=row, column=col).border = thin_border
    _auto_fit(ws)

    ws2 = wb.create_sheet("Servicios")
    headers = ["ID", "Cliente", "Origen", "Destino", "Valor", "Pagado", "Saldo", "Estado", "Recogido", "Entregado"]
    ws2.append(headers)
    for idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    for servicio in servicios:
        ws2.append(
            [
                _xls(servicio.id),
                _xls(getattr(servicio.cliente, "nombre", "")),
                _xls(servicio.origen or "-"),
                _xls(servicio.destino or "-"),
                _xls(servicio.valor or 0),
                _xls(servicio.total_pagado or 0),
                _xls(servicio.saldo_cartera),
                _xls(servicio.get_estado_pago_display()),
                _xls(servicio.recogido_en),
                _xls(servicio.entregado_en),
            ]
        )

    for row in range(2, ws2.max_row + 1):
        for col in (5, 6, 7):
            _money_fmt(ws2.cell(row=row, column=col)).alignment = right
        for col in (9, 10):
            ws2.cell(row=row, column=col).number_format = "yyyy-mm-dd hh:mm"
    for row in range(1, ws2.max_row + 1):
        for col in range(1, ws2.max_column + 1):
            ws2.cell(row=row, column=col).border = thin_border
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}1"
    _auto_fit(ws2)

    ws3 = wb.create_sheet("Notas")
    ws3["A1"] = "Generado por"
    ws3["B1"] = _xls(request.user.get_username())
    ws3["A2"] = "Empresa"
    ws3["B2"] = _xls(getattr(ruta.empresa, "nombre", str(ruta.empresa)))
    ws3["A3"] = "Nota"
    ws3["B3"] = "Exportar no modifica el estado de la ruta."
    for cell in ("A1", "A2", "A3"):
        ws3[cell].font = bold
    _auto_fit(ws3)

    filename = f"Cierre_ruta_{ruta.id}_{ruta.fecha_salida}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ReordenarServiciosView(View):
    def post(self, request, ruta_id):
        if not request.user.is_authenticated:
            return HttpResponseForbidden("Auth requerida")
        if not is_gerente(request.user):
            return HttpResponseForbidden("Solo gerente")

        ruta = _ruta_autorizada(ruta_id, request.user, gerente_only=True)
        if ruta is None:
            return HttpResponseForbidden("No autorizado")
        if ruta.estado != "ACTIVA":
            return HttpResponseBadRequest("No se puede reordenar una ruta cerrada")

        try:
            data = json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return HttpResponseBadRequest("JSON invalido")
        order = data if isinstance(data, list) else data.get("order")
        if not isinstance(order, list) or not all(isinstance(item, int) for item in order):
            return HttpResponseBadRequest("Formato de order invalido")
        if len(order) != len(set(order)):
            return HttpResponseBadRequest("IDs duplicados")

        servicios = list(Servicio.objects.filter(ruta=ruta).order_by("orden", "id"))
        ids_ruta = {servicio.id for servicio in servicios}
        if set(order) != ids_ruta:
            return HttpResponseBadRequest("IDs no coinciden con la ruta")

        by_id = {servicio.id: servicio for servicio in servicios}
        with transaction.atomic():
            for pos, sid in enumerate(order, start=1):
                by_id[sid].orden = pos
            Servicio.objects.bulk_update(servicios, ["orden"])
            send_webpush_to_empresa(
                ruta.empresa,
                "Orden de ruta actualizado",
                f"Ruta #{ruta.id}: se modifico el orden de servicios.",
                {"url": f"/rutas/{ruta.id}/hoja/"},
                exclude_user=request.user,
            )
        return JsonResponse({"ok": True})


@login_required
def por_ruta(request, ruta_id: int):
    return redirect("servicios:por_ruta", ruta_id=ruta_id)
